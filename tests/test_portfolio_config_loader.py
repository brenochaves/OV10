from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from ov10.config import load_portfolio_book_config

REFERENCE_WORKBOOK_PATH = next(Path("ov10_codex_handoff").glob("*.xlsx"))
FALLBACK_CONFIG_PATH = Path("config/ov10_portfolio.toml")


class PortfolioConfigLoaderTests(unittest.TestCase):
    def test_toml_config_rejects_invalid_instrument_types_shape(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            config_path = Path(temp_dir) / "invalid_config.toml"
            config_path.write_text(
                """
[metadata]
default_portfolio_id = "core"

[[portfolios]]
portfolio_id = "core"
name = "Core"
short_name = "core"

[[books]]
book_id = "stocks"
portfolio_id = "core"
name = "Stocks"
instrument_types = "stock_br"

[[account_assignments]]
rule_id = "all_accounts"
portfolio_id = "core"
account_patterns = ["*"]
""".strip(),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(
                ValueError,
                "books\\[0\\]\\.instrument_types must be a non-empty array of strings",
            ):
                load_portfolio_book_config(config_path)

    def test_workbook_config_loads_named_portfolio_and_supported_books(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "portfolio_config.xlsx"
            workbook = Workbook()
            workbook.remove(workbook.active)

            portfolios_ws = workbook.create_sheet("config.carteiras")
            portfolios_ws["A9"] = 1
            portfolios_ws["B9"] = True
            portfolios_ws["C9"] = "Long Term"
            portfolios_ws["D9"] = 0.75

            books_ws = workbook.create_sheet("config.books")
            books_ws["A6"] = 1
            books_ws["B6"] = "Renda Fixa"
            books_ws["C6"] = "Tesouro Direto"
            books_ws["D6"] = 0.2
            books_ws["A7"] = 2
            books_ws["B7"] = "Renda Variável"
            books_ws["C7"] = "Ações"
            books_ws["D7"] = 0.5
            books_ws["E7"] = "AÇÃO"
            books_ws["A8"] = 3
            books_ws["B8"] = "Renda Variável"
            books_ws["C8"] = "Stocks"
            books_ws["D8"] = 0.25
            books_ws["E8"] = "STOCK"
            books_ws["A9"] = 4
            books_ws["B9"] = "Renda Variável"
            books_ws["C9"] = "Proteções"
            books_ws["D9"] = 0.05
            books_ws["E9"] = "OPÇÃO"

            workbook.save(workbook_path)

            config = load_portfolio_book_config(workbook_path)

        self.assertEqual(config.default_portfolio_id, "long_term")
        self.assertEqual(len(config.portfolios), 1)
        self.assertEqual(config.portfolios[0].portfolio_id, "long_term")
        self.assertEqual(config.account_assignments[0].portfolio_id, "long_term")

        books_by_id = {item.book_id: item for item in config.books}
        self.assertEqual(
            set(books_by_id),
            {"tesouro_direto", "acoes", "stocks", "protecoes", "sem_book"},
        )
        self.assertTrue(books_by_id["tesouro_direto"].is_active)
        self.assertTrue(books_by_id["acoes"].is_active)
        self.assertTrue(books_by_id["stocks"].is_active)
        self.assertFalse(books_by_id["protecoes"].is_active)
        self.assertTrue(books_by_id["sem_book"].is_active)

    def test_reference_workbook_uses_fallback_portfolio_config_when_template_is_generic(
        self,
    ) -> None:
        config = load_portfolio_book_config(
            REFERENCE_WORKBOOK_PATH,
            fallback_path=FALLBACK_CONFIG_PATH,
        )

        self.assertEqual(config.default_portfolio_id, "core")
        self.assertEqual(len(config.portfolios), 1)
        self.assertEqual(config.portfolios[0].portfolio_id, "core")
        self.assertEqual(config.account_assignments[0].portfolio_id, "core")

        books_by_id = {item.book_id: item for item in config.books}
        self.assertEqual(len(books_by_id), 11)
        self.assertEqual(
            {book_id for book_id, item in books_by_id.items() if item.is_active},
            {"fundos", "tesouro_direto", "acoes", "fiis", "stocks", "criptos", "sem_book"},
        )
