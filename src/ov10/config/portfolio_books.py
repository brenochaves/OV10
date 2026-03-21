from __future__ import annotations

import re
import tomllib
import unicodedata
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

from ov10.domain.enums import InstrumentType
from ov10.domain.models import BookDefinition, PortfolioDefinition

WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
GENERIC_PORTFOLIO_RE = re.compile(r"^carteira #\d+$", re.IGNORECASE)


class _CellLike(Protocol):
    @property
    def value(self) -> object: ...


class _WorksheetLike(Protocol):
    def cell(self, row: int, column: int) -> _CellLike: ...


def _require_non_empty(value: str, field_name: str) -> str:
    normalized = value.strip()
    if not normalized:
        raise ValueError(f"{field_name} cannot be empty")
    return normalized


@dataclass(frozen=True, slots=True)
class AccountAssignmentRule:
    rule_id: str
    portfolio_id: str
    account_patterns: tuple[str, ...]

    def __post_init__(self) -> None:
        object.__setattr__(self, "rule_id", _require_non_empty(self.rule_id, "rule_id"))
        object.__setattr__(
            self,
            "portfolio_id",
            _require_non_empty(self.portfolio_id, "portfolio_id"),
        )
        if not self.account_patterns:
            raise ValueError("account_patterns cannot be empty")


@dataclass(frozen=True, slots=True)
class PortfolioBookConfig:
    source_file: str
    default_portfolio_id: str
    portfolios: tuple[PortfolioDefinition, ...]
    books: tuple[BookDefinition, ...]
    account_assignments: tuple[AccountAssignmentRule, ...]

    def __post_init__(self) -> None:
        object.__setattr__(self, "source_file", _require_non_empty(self.source_file, "source_file"))
        object.__setattr__(
            self,
            "default_portfolio_id",
            _require_non_empty(self.default_portfolio_id, "default_portfolio_id"),
        )
        if not self.portfolios:
            raise ValueError("portfolios cannot be empty")
        if not self.books:
            raise ValueError("books cannot be empty")
        portfolio_ids = {item.portfolio_id for item in self.portfolios}
        if self.default_portfolio_id not in portfolio_ids:
            raise ValueError("default_portfolio_id must reference an existing portfolio")
        invalid_book_portfolios = sorted(
            {item.portfolio_id for item in self.books if item.portfolio_id not in portfolio_ids}
        )
        if invalid_book_portfolios:
            raise ValueError("books must reference existing portfolios")
        invalid_assignment_portfolios = sorted(
            {
                item.portfolio_id
                for item in self.account_assignments
                if item.portfolio_id not in portfolio_ids
            }
        )
        if invalid_assignment_portfolios:
            raise ValueError("account assignments must reference existing portfolios")


@dataclass(frozen=True, slots=True)
class _WorkbookPortfolioCandidate:
    row_number: int
    name: str
    weight: Decimal


@dataclass(frozen=True, slots=True)
class _WorkbookBookTemplate:
    book_slug: str
    name: str
    target_weight: Decimal
    instrument_types: tuple[InstrumentType, ...]
    is_active: bool
    sort_order: int


def load_portfolio_book_config(
    path: str | Path,
    *,
    fallback_path: str | Path | None = None,
) -> PortfolioBookConfig:
    config_path = Path(path)
    suffix = config_path.suffix.lower()

    if suffix == ".toml":
        return _load_toml_portfolio_book_config(config_path)

    if suffix in WORKBOOK_SUFFIXES:
        fallback_config: PortfolioBookConfig | None = None
        if fallback_path is not None:
            fallback_candidate = Path(fallback_path)
            if fallback_candidate.resolve() != config_path.resolve():
                fallback_config = load_portfolio_book_config(fallback_candidate)
        return _load_workbook_portfolio_book_config(config_path, fallback_config=fallback_config)

    raise ValueError(f"Unsupported config file type: {config_path.suffix}")


def _load_toml_portfolio_book_config(path: Path) -> PortfolioBookConfig:
    raw = tomllib.loads(path.read_text(encoding="utf-8"))
    _validate_toml_portfolio_book_config(raw)

    metadata = raw.get("metadata", {})
    portfolios = tuple(_load_portfolios(raw.get("portfolios", [])))
    books = tuple(_load_books(raw.get("books", [])))
    assignments = tuple(_load_account_assignments(raw.get("account_assignments", [])))

    return PortfolioBookConfig(
        source_file=str(path.resolve()),
        default_portfolio_id=str(metadata["default_portfolio_id"]),
        portfolios=portfolios,
        books=books,
        account_assignments=assignments,
    )


def _load_workbook_portfolio_book_config(
    path: Path,
    *,
    fallback_config: PortfolioBookConfig | None,
) -> PortfolioBookConfig:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if "config.carteiras" not in workbook.sheetnames:
            raise ValueError("Workbook config is missing sheet `config.carteiras`")
        if "config.books" not in workbook.sheetnames:
            raise ValueError("Workbook config is missing sheet `config.books`")

        portfolio_candidates = _load_workbook_portfolio_candidates(workbook["config.carteiras"])
        portfolio_definitions = _materialize_workbook_portfolios(portfolio_candidates)
        if portfolio_definitions:
            portfolios = portfolio_definitions
            default_portfolio_id = portfolios[0].portfolio_id
            account_assignments = (
                AccountAssignmentRule(
                    rule_id=f"all_accounts_to_{default_portfolio_id}",
                    portfolio_id=default_portfolio_id,
                    account_patterns=("*",),
                ),
            )
        elif fallback_config is not None:
            portfolios = fallback_config.portfolios
            default_portfolio_id = fallback_config.default_portfolio_id
            account_assignments = fallback_config.account_assignments
        else:
            portfolios = (
                PortfolioDefinition(
                    portfolio_id="core",
                    name="Core Portfolio",
                    short_name="core",
                    profile="workbook_fallback",
                    is_active=True,
                ),
            )
            default_portfolio_id = "core"
            account_assignments = (
                AccountAssignmentRule(
                    rule_id="all_accounts_to_core",
                    portfolio_id="core",
                    account_patterns=("*",),
                ),
            )

        book_templates = _load_workbook_book_templates(workbook["config.books"])
        if not book_templates and fallback_config is not None:
            books = fallback_config.books
        else:
            books = tuple(_expand_book_templates(book_templates, portfolios))

        return PortfolioBookConfig(
            source_file=str(path.resolve()),
            default_portfolio_id=default_portfolio_id,
            portfolios=portfolios,
            books=books,
            account_assignments=account_assignments,
        )
    finally:
        workbook.close()


def _validate_toml_portfolio_book_config(raw: object) -> None:
    document = _require_mapping(raw, "portfolio config document")
    metadata = _require_mapping(document.get("metadata"), "metadata")
    _require_string(metadata.get("default_portfolio_id"), "metadata.default_portfolio_id")

    for index, item in enumerate(_require_mapping_list(document.get("portfolios"), "portfolios")):
        context = f"portfolios[{index}]"
        _require_string(item.get("portfolio_id"), f"{context}.portfolio_id")
        _require_string(item.get("name"), f"{context}.name")
        _require_string(item.get("short_name"), f"{context}.short_name")
        if "profile" in item:
            _require_string(item.get("profile"), f"{context}.profile")
        if "is_active" in item:
            _require_bool(item.get("is_active"), f"{context}.is_active")

    for index, item in enumerate(_require_mapping_list(document.get("books"), "books")):
        context = f"books[{index}]"
        _require_string(item.get("book_id"), f"{context}.book_id")
        _require_string(item.get("portfolio_id"), f"{context}.portfolio_id")
        _require_string(item.get("name"), f"{context}.name")
        if "target_weight" in item:
            _coerce_decimal(item.get("target_weight"), f"{context}.target_weight")
        _require_string_sequence(item.get("instrument_types"), f"{context}.instrument_types")
        if "is_active" in item:
            _require_bool(item.get("is_active"), f"{context}.is_active")
        if "sort_order" in item:
            _coerce_int_like(item.get("sort_order"), f"{context}.sort_order")

    for index, item in enumerate(
        _require_mapping_list(document.get("account_assignments"), "account_assignments")
    ):
        context = f"account_assignments[{index}]"
        _require_string(item.get("rule_id"), f"{context}.rule_id")
        _require_string(item.get("portfolio_id"), f"{context}.portfolio_id")
        _require_string_sequence(item.get("account_patterns"), f"{context}.account_patterns")


def _load_portfolios(raw_portfolios: object) -> list[PortfolioDefinition]:
    validated = _require_mapping_list(raw_portfolios, "portfolios")
    return [
        PortfolioDefinition(
            portfolio_id=str(item["portfolio_id"]),
            name=str(item["name"]),
            short_name=str(item["short_name"]),
            profile=str(item.get("profile", "custom")),
            is_active=bool(item.get("is_active", True)),
        )
        for item in validated
    ]


def _load_books(raw_books: object) -> list[BookDefinition]:
    validated = _require_mapping_list(raw_books, "books")
    books: list[BookDefinition] = []
    for item in validated:
        books.append(
            BookDefinition(
                book_id=str(item["book_id"]),
                portfolio_id=str(item["portfolio_id"]),
                name=str(item["name"]),
                target_weight=_coerce_decimal(
                    item.get("target_weight", "0"), "books.target_weight"
                ),
                instrument_types=tuple(
                    InstrumentType(value)
                    for value in _require_string_sequence(
                        item.get("instrument_types"),
                        "books.instrument_types",
                    )
                ),
                is_active=bool(item.get("is_active", True)),
                sort_order=_coerce_int_like(item.get("sort_order", 0), "books.sort_order"),
            )
        )
    return books


def _load_account_assignments(
    raw_rules: object,
) -> list[AccountAssignmentRule]:
    validated = _require_mapping_list(raw_rules, "account_assignments")
    return [
        AccountAssignmentRule(
            rule_id=str(item["rule_id"]),
            portfolio_id=str(item["portfolio_id"]),
            account_patterns=_require_string_sequence(
                item.get("account_patterns"),
                "account_assignments.account_patterns",
            ),
        )
        for item in validated
    ]


def _load_workbook_portfolio_candidates(
    worksheet: _WorksheetLike,
) -> list[_WorkbookPortfolioCandidate]:
    candidates: list[_WorkbookPortfolioCandidate] = []
    for row_number in range(9, 24):
        active = worksheet.cell(row=row_number, column=2).value
        name = _as_clean_text(worksheet.cell(row=row_number, column=3).value)
        if not _is_truthy(active) or not name:
            continue
        candidates.append(
            _WorkbookPortfolioCandidate(
                row_number=row_number,
                name=name,
                weight=_as_decimal(worksheet.cell(row=row_number, column=4).value),
            )
        )
    return candidates


def _materialize_workbook_portfolios(
    candidates: list[_WorkbookPortfolioCandidate],
) -> tuple[PortfolioDefinition, ...]:
    if not candidates:
        return ()

    if not any(
        candidate.weight > Decimal("0") or not _is_generic_portfolio_name(candidate.name)
        for candidate in candidates
    ):
        return ()

    seen_ids: set[str] = set()
    default_candidate = sorted(
        candidates,
        key=lambda item: (-item.weight, item.row_number),
    )[0]
    definitions: list[PortfolioDefinition] = []
    for candidate in candidates:
        fallback_slug = f"portfolio_{candidate.row_number}"
        portfolio_id = _make_unique_slug(candidate.name, seen_ids, fallback_slug)
        profile = (
            "workbook_default"
            if candidate.row_number == default_candidate.row_number
            else "workbook_import"
        )
        definitions.append(
            PortfolioDefinition(
                portfolio_id=portfolio_id,
                name=candidate.name,
                short_name=portfolio_id,
                profile=profile,
                is_active=True,
            )
        )
    default_portfolio_slug = _slugify(default_candidate.name)
    definitions.sort(
        key=lambda item: (
            item.portfolio_id != default_portfolio_slug,
            item.portfolio_id,
        )
    )
    return tuple(definitions)


def _load_workbook_book_templates(worksheet: _WorksheetLike) -> tuple[_WorkbookBookTemplate, ...]:
    templates: list[_WorkbookBookTemplate] = []
    seen_slugs: set[str] = set()

    for row_number in range(6, 120):
        order_value = worksheet.cell(row=row_number, column=1).value
        strategy = _as_clean_text(worksheet.cell(row=row_number, column=2).value)
        name = _as_clean_text(worksheet.cell(row=row_number, column=3).value)
        weight_value = worksheet.cell(row=row_number, column=4).value
        class_name = _as_clean_text(worksheet.cell(row=row_number, column=5).value)

        if order_value in (None, "") or not strategy or not name:
            continue

        sort_order = int(Decimal(str(order_value)))
        fallback_slug = f"book_{sort_order}"
        book_slug = _make_unique_slug(name, seen_slugs, fallback_slug)
        instrument_types, is_active = _infer_book_instrument_types(name=name, class_name=class_name)
        templates.append(
            _WorkbookBookTemplate(
                book_slug=book_slug,
                name=name,
                target_weight=_as_decimal(weight_value),
                instrument_types=instrument_types,
                is_active=is_active,
                sort_order=sort_order,
            )
        )

    if templates:
        templates.append(
            _WorkbookBookTemplate(
                book_slug=_make_unique_slug("sem_book", seen_slugs, "sem_book"),
                name="Sem Book",
                target_weight=Decimal("0"),
                instrument_types=(InstrumentType.UNKNOWN,),
                is_active=True,
                sort_order=max(item.sort_order for item in templates) + 1,
            )
        )

    return tuple(templates)


def _expand_book_templates(
    templates: tuple[_WorkbookBookTemplate, ...],
    portfolios: tuple[PortfolioDefinition, ...],
) -> list[BookDefinition]:
    if not templates:
        return []

    single_portfolio = len(portfolios) == 1
    books: list[BookDefinition] = []
    for portfolio in portfolios:
        for template in templates:
            if single_portfolio:
                book_id = template.book_slug
            else:
                book_id = f"{portfolio.portfolio_id}__{template.book_slug}"
            books.append(
                BookDefinition(
                    book_id=book_id,
                    portfolio_id=portfolio.portfolio_id,
                    name=template.name,
                    target_weight=template.target_weight,
                    instrument_types=template.instrument_types,
                    is_active=template.is_active,
                    sort_order=template.sort_order,
                )
            )
    return books


def _infer_book_instrument_types(
    *,
    name: str,
    class_name: str | None,
) -> tuple[tuple[InstrumentType, ...], bool]:
    normalized_name = _normalize_label(name)
    normalized_class = _normalize_label(class_name)

    if normalized_class == "ACAO" or normalized_name == "ACOES":
        return (
            InstrumentType.STOCK_BR,
            InstrumentType.BDR,
            InstrumentType.ETF_BR,
        ), True
    if normalized_class == "FII" or normalized_name == "FIIS":
        return (InstrumentType.FII,), True
    if normalized_class == "STOCK" or normalized_name == "STOCKS":
        return (
            InstrumentType.STOCK_US,
            InstrumentType.ETF_US,
        ), True
    if normalized_class == "CRIPTO" or normalized_name == "CRIPTOS":
        return (InstrumentType.CRYPTO,), True
    if normalized_name == "TESOURO DIRETO":
        return (InstrumentType.TREASURY_BR,), True
    if normalized_name == "FUNDOS":
        return (InstrumentType.FUND_BR,), True

    return (InstrumentType.UNKNOWN,), False


def _as_clean_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    return str(value).strip() or None


def _as_decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, bool):
        return Decimal("1") if value else Decimal("0")
    return Decimal(str(value))


def _is_truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int | float | Decimal):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "sim", "yes"}
    return False


def _is_generic_portfolio_name(name: str) -> bool:
    return GENERIC_PORTFOLIO_RE.match(name.strip()) is not None


def _normalize_label(value: str | None) -> str:
    if value is None:
        return ""
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    collapsed = re.sub(r"[^A-Za-z0-9]+", " ", ascii_only)
    return collapsed.strip().upper()


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    collapsed = re.sub(r"[^A-Za-z0-9]+", "_", ascii_only)
    slug = collapsed.strip("_").lower()
    return slug


def _make_unique_slug(value: str, seen_ids: set[str], fallback: str) -> str:
    base = _slugify(value) or fallback
    candidate = base
    index = 2
    while candidate in seen_ids:
        candidate = f"{base}_{index}"
        index += 1
    seen_ids.add(candidate)
    return candidate


def _require_mapping(value: object, field_name: str) -> dict[str, object]:
    if not isinstance(value, Mapping):
        raise ValueError(f"{field_name} must be a TOML table")
    return {str(key): item for key, item in value.items()}


def _require_mapping_list(value: object, field_name: str) -> list[dict[str, object]]:
    if value is None:
        return []
    if not isinstance(value, list):
        raise ValueError(f"{field_name} must be an array of tables")
    items: list[dict[str, object]] = []
    for index, item in enumerate(value):
        if not isinstance(item, Mapping):
            raise ValueError(f"{field_name}[{index}] must be a TOML table")
        items.append({str(key): entry for key, entry in item.items()})
    return items


def _require_string(value: object, field_name: str) -> str:
    if not isinstance(value, str):
        raise ValueError(f"{field_name} must be a string")
    return _require_non_empty(value, field_name)


def _require_bool(value: object, field_name: str) -> bool:
    if not isinstance(value, bool):
        raise ValueError(f"{field_name} must be a boolean")
    return value


def _require_string_sequence(value: object, field_name: str) -> tuple[str, ...]:
    if not isinstance(value, Sequence) or isinstance(value, str | bytes):
        raise ValueError(f"{field_name} must be a non-empty array of strings")
    normalized = tuple(_require_string(item, f"{field_name}[]") for item in value)
    if not normalized:
        raise ValueError(f"{field_name} must not be empty")
    return normalized


def _coerce_decimal(value: object, field_name: str) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"{field_name} must be decimal-compatible") from exc


def _coerce_int_like(value: object, field_name: str) -> int:
    try:
        return int(Decimal(str(value)))
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"{field_name} must be int-compatible") from exc
