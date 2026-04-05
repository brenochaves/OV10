from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook

from ov10.allocation import assign_account_positions_to_books, assign_accounts_to_portfolios
from ov10.cash import build_cash_ledger
from ov10.config import load_portfolio_book_config
from ov10.domain.models import BookPositionSnapshot, CanonicalDividendReceipt, PositionSnapshot
from ov10.ingestion import load_status_invest_workbook
from ov10.positions import compute_account_positions, compute_positions
from ov10.reconciliation.market_valuation import (
    MarketValuationContext,
    build_market_valuation_context,
)
from ov10.storage import SQLiteOV10Store

DEFAULT_REFERENCE_WORKBOOK_PATH = (
    Path("ov10_codex_handoff") / "Cópia de v6.4_Controle de Investimentos (Alpha).xlsx"
)
DEFAULT_CONFIG_PATH = Path("config/ov10_portfolio.toml")
DEFAULT_DATABASE_PATH = Path("var/ov10.sqlite3")


class _CellLike(Protocol):
    @property
    def value(self) -> object: ...


class _WorksheetLike(Protocol):
    def cell(self, row: int, column: int) -> _CellLike: ...


@dataclass(frozen=True, slots=True)
class ReferenceGap:
    reference_field: str
    support_level: str
    classification: str
    current_source: str
    note: str


@dataclass(frozen=True, slots=True)
class ReferenceAreaReport:
    area_id: str
    sheet_name: str
    reference_range: str
    overall_classification: str
    summary: str
    gaps: tuple[ReferenceGap, ...]
    metrics: dict[str, object]

    def to_dict(self) -> dict[str, object]:
        payload = asdict(self)
        payload["metrics"] = _serialize_value(self.metrics)
        return payload


@dataclass(frozen=True, slots=True)
class ReferenceWorkbookReconciliationReport:
    source_workbook: str
    reference_workbook: str
    config_path: str
    batch_id: str
    transaction_count: int
    dividend_receipt_count: int
    corporate_action_count: int
    snapshot_count: int
    account_snapshot_count: int
    book_snapshot_count: int
    cash_movement_count: int
    cash_balance_snapshot_count: int
    areas: tuple[ReferenceAreaReport, ...]

    def to_dict(self) -> dict[str, object]:
        return {
            "source_workbook": self.source_workbook,
            "reference_workbook": self.reference_workbook,
            "config_path": self.config_path,
            "batch_id": self.batch_id,
            "transaction_count": self.transaction_count,
            "dividend_receipt_count": self.dividend_receipt_count,
            "corporate_action_count": self.corporate_action_count,
            "snapshot_count": self.snapshot_count,
            "account_snapshot_count": self.account_snapshot_count,
            "book_snapshot_count": self.book_snapshot_count,
            "cash_movement_count": self.cash_movement_count,
            "cash_balance_snapshot_count": self.cash_balance_snapshot_count,
            "areas": [item.to_dict() for item in self.areas],
            "summary": {
                "expected_areas": sum(
                    1 for item in self.areas if item.overall_classification == "expected"
                ),
                "explained_areas": sum(
                    1 for item in self.areas if item.overall_classification == "explained"
                ),
                "blocking_areas": sum(
                    1 for item in self.areas if item.overall_classification == "blocking"
                ),
            },
        }


@dataclass(frozen=True, slots=True)
class _RuntimeContext:
    batch_id: str
    transaction_count: int
    dividend_receipt_count: int
    corporate_action_count: int
    snapshot_count: int
    account_snapshot_count: int
    book_snapshot_count: int
    cash_movement_count: int
    cash_balance_snapshot_count: int
    last_snapshot_date: date | None
    dividend_receipts: tuple[CanonicalDividendReceipt, ...]
    book_ids: tuple[str, ...]
    market_valuation: MarketValuationContext


def generate_reference_workbook_reconciliation(
    source_workbook_path: str | Path,
    *,
    reference_workbook_path: str | Path = DEFAULT_REFERENCE_WORKBOOK_PATH,
    config_path: str | Path = DEFAULT_CONFIG_PATH,
    database_path: str | Path | None = None,
) -> ReferenceWorkbookReconciliationReport:
    source_workbook_path = Path(source_workbook_path)
    reference_workbook_path = Path(reference_workbook_path)
    config_path = Path(config_path)
    context = _build_runtime_context(
        source_workbook_path,
        config_path=config_path,
        database_path=None if database_path is None else Path(database_path),
    )

    workbook_formula = load_workbook(reference_workbook_path, data_only=False, read_only=True)
    workbook_values = load_workbook(reference_workbook_path, data_only=True, read_only=True)
    try:
        areas = (
            _build_portfolio_area(workbook_values["portfolio."], context),
            _build_cash_area(workbook_values["caixa"], context),
            _build_allocation_area(workbook_values["alocação"], context),
            _build_income_area(
                worksheet_formulas=workbook_formula["rendimentos"],
                worksheet_values=workbook_values["rendimentos"],
                context=context,
            ),
        )
    finally:
        workbook_formula.close()
        workbook_values.close()

    return ReferenceWorkbookReconciliationReport(
        source_workbook=str(source_workbook_path.resolve()),
        reference_workbook=str(reference_workbook_path.resolve()),
        config_path=str(config_path.resolve()),
        batch_id=context.batch_id,
        transaction_count=context.transaction_count,
        dividend_receipt_count=context.dividend_receipt_count,
        corporate_action_count=context.corporate_action_count,
        snapshot_count=context.snapshot_count,
        account_snapshot_count=context.account_snapshot_count,
        book_snapshot_count=context.book_snapshot_count,
        cash_movement_count=context.cash_movement_count,
        cash_balance_snapshot_count=context.cash_balance_snapshot_count,
        areas=areas,
    )


def _build_runtime_context(
    source_workbook_path: Path,
    *,
    config_path: Path,
    database_path: Path | None,
) -> _RuntimeContext:
    bundle = load_status_invest_workbook(source_workbook_path)
    cash_ledger = build_cash_ledger(bundle)
    fallback_config_path = (
        DEFAULT_CONFIG_PATH
        if config_path.suffix.lower() in {".xlsx", ".xlsm"}
        and config_path.resolve() != DEFAULT_CONFIG_PATH.resolve()
        else None
    )
    portfolio_config = load_portfolio_book_config(config_path, fallback_path=fallback_config_path)
    account_ids_by_broker = {item.account_name: item.account_id for item in cash_ledger.accounts}
    positions = compute_positions(bundle.transactions)
    account_positions = compute_account_positions(
        bundle.transactions,
        account_ids_by_broker=account_ids_by_broker,
    )
    account_assignments = assign_accounts_to_portfolios(
        cash_ledger.accounts,
        portfolio_config,
        batch_id=bundle.batch.batch_id,
    )
    book_positions = assign_account_positions_to_books(
        account_positions,
        account_assignments,
        portfolio_config,
    )
    last_snapshot_date = max(
        (item.snapshot_date for item in positions),
        default=None,
    )
    market_valuation = _load_market_valuation(
        positions=positions,
        book_positions=book_positions,
        database_path=database_path,
        snapshot_date=last_snapshot_date,
    )
    return _RuntimeContext(
        batch_id=bundle.batch.batch_id,
        transaction_count=len(bundle.transactions),
        dividend_receipt_count=len(bundle.dividend_receipts),
        corporate_action_count=len(bundle.corporate_actions),
        snapshot_count=len(positions),
        account_snapshot_count=len(account_positions),
        book_snapshot_count=len(book_positions),
        cash_movement_count=len(cash_ledger.movements),
        cash_balance_snapshot_count=len(cash_ledger.balance_snapshots),
        last_snapshot_date=last_snapshot_date,
        dividend_receipts=bundle.dividend_receipts,
        book_ids=tuple(sorted({item.book_id for item in book_positions})),
        market_valuation=market_valuation,
    )


def _build_portfolio_area(
    worksheet: _WorksheetLike, context: _RuntimeContext
) -> ReferenceAreaReport:
    headers = _clean_headers(
        worksheet=worksheet,
        row=3,
        start_column=1,
        end_column=25,
    )
    support = {
        "ativo": _gap(
            "ativo",
            "matched",
            "expected",
            "position_snapshots.instrument_code",
            "O codigo do ativo ja existe no snapshot canonico.",
        ),
        "market_cod": _coverage_gap(
            "market_cod",
            matched_count=context.market_valuation.positions_with_market_code,
            total_count=context.market_valuation.position_count,
            current_source="latest market_snapshot.market",
            matched_note=(
                "O codigo de mercado ja pode ser lido do snapshot persistido mais recente."
            ),
            partial_note=(
                "Existe cobertura parcial de codigo de mercado a partir dos snapshots persistidos."
            ),
            missing_note=(
                "Nenhum snapshot persistido com mercado resolvido foi encontrado para "
                "as posicoes abertas."
            ),
        ),
        "nome_pregao": _coverage_gap(
            "nome_pregao",
            matched_count=context.market_valuation.positions_with_display_name,
            total_count=context.market_valuation.position_count,
            current_source="latest market_snapshot.provider_metadata.longName",
            matched_note=("O nome de pregão ja pode ser lido da metadata persistida do provider."),
            partial_note=(
                "Existe cobertura parcial de nome de pregão a partir da metadata persistida."
            ),
            missing_note=(
                "Nenhum nome de pregão persistido foi encontrado para as posicoes abertas."
            ),
        ),
        "classe": _gap(
            "classe",
            "partial",
            "explained",
            "position_snapshots.instrument_type",
            (
                "O OV10 tem tipo canônico de instrumento, "
                "mas ainda nao o rotulo de classe da planilha."
            ),
        ),
        "corretora": _gap(
            "corretora",
            "matched",
            "expected",
            "account_position_snapshots.account_id/account_name",
            "A granularidade por corretora existe nos snapshots por conta.",
        ),
        "investidor": _gap(
            "investidor",
            "missing",
            "blocking",
            "",
            "Ainda nao existe dimensao de investidor separada de conta/carteira.",
        ),
        "setor": _gap(
            "setor",
            "missing",
            "blocking",
            "",
            "Depende de referencia governada de instrumento.",
        ),
        "subsetor": _gap(
            "subsetor",
            "missing",
            "blocking",
            "",
            "Depende de referencia governada de instrumento.",
        ),
        "date_ini": _gap(
            "date_ini",
            "partial",
            "explained",
            "canonical_transaction.trade_date",
            (
                "A data inicial pode ser derivada do primeiro negocio, "
                "mas nao esta exposta no report atual."
            ),
        ),
        "qtd_atual": _gap(
            "qtd_atual",
            "matched",
            "expected",
            "position_snapshots.quantity",
            "Quantidade atual ja existe nos snapshots agregados e por conta.",
        ),
        "pm_atual": _gap(
            "pm_atual",
            "matched",
            "expected",
            "position_snapshots.avg_cost",
            "Preco medio atual ja existe nos snapshots canonicos.",
        ),
        "pm_fx": _coverage_gap(
            "pm_fx",
            matched_count=context.market_valuation.positions_with_avg_cost_in_price_currency,
            total_count=context.market_valuation.position_count,
            current_source="position_snapshots.avg_cost + latest fx_snapshot",
            matched_note=(
                "O preco medio em moeda da cotacao ja pode ser derivado da camada persistida."
            ),
            partial_note=(
                "Existe cobertura parcial de traducao do preco medio para a moeda de cotacao."
            ),
            missing_note=(
                "Nenhuma traducao persistida de preco medio para moeda de cotacao foi encontrada."
            ),
        ),
        "pm_ptax": _gap(
            "pm_ptax",
            "missing",
            "blocking",
            "",
            "Nao existe saida de preco medio ajustado por PTAX.",
        ),
        "fator_pts": _gap(
            "fator_pts",
            "missing",
            "blocking",
            "",
            "O contrato atual nao possui fator de pontos/conversao exposto no report.",
        ),
        "vlr_investido": _gap(
            "vlr_investido",
            "matched",
            "expected",
            "position_snapshots.total_cost",
            "Valor investido ja existe como custo total nos snapshots.",
        ),
        "price": _coverage_gap(
            "price",
            matched_count=context.market_valuation.positions_with_market_price,
            total_count=context.market_valuation.position_count,
            current_source="latest market_snapshot.price",
            matched_note=(
                "O preco de mercado ja pode ser lido do snapshot persistido mais recente."
            ),
            partial_note=(
                "Existe cobertura parcial de preco de mercado a partir dos snapshots persistidos."
            ),
            missing_note=(
                "Nenhum preco de mercado persistido foi encontrado para as posicoes abertas."
            ),
        ),
        "change": _coverage_gap(
            "change",
            matched_count=context.market_valuation.positions_with_change,
            total_count=context.market_valuation.position_count,
            current_source="latest market_snapshot.absolute_change",
            matched_note=(
                "A variacao absoluta ja pode ser lida do snapshot persistido mais recente."
            ),
            partial_note=(
                "Existe cobertura parcial de variacao absoluta a partir dos snapshots persistidos."
            ),
            missing_note=(
                "Nenhuma variacao absoluta persistida foi encontrada para as posicoes abertas."
            ),
        ),
        "changepct": _coverage_gap(
            "changepct",
            matched_count=context.market_valuation.positions_with_change_pct,
            total_count=context.market_valuation.position_count,
            current_source="latest market_snapshot.percent_change",
            matched_note=(
                "A variacao percentual ja pode ser lida do snapshot persistido mais recente."
            ),
            partial_note=(
                "Existe cobertura parcial de variacao percentual a partir dos snapshots "
                "persistidos."
            ),
            missing_note=(
                "Nenhuma variacao percentual persistida foi encontrada para as posicoes abertas."
            ),
        ),
        "moeda_price": _coverage_gap(
            "moeda_price",
            matched_count=context.market_valuation.positions_with_price_currency,
            total_count=context.market_valuation.position_count,
            current_source="latest market_snapshot.currency",
            matched_note="A moeda da cotacao ja existe no snapshot persistido de mercado.",
            partial_note=(
                "Existe cobertura parcial de moeda da cotacao a partir dos snapshots persistidos."
            ),
            missing_note=(
                "Nenhuma moeda de cotacao persistida foi encontrada para as posicoes abertas."
            ),
        ),
        "corretora_investidor": _gap(
            "corretora_investidor",
            "partial",
            "explained",
            "account_position_snapshots + account_portfolio_assignment",
            "O join de corretora e carteira existe, mas nao ha eixo separado de investidor.",
        ),
        "categoria_subtotal": _gap(
            "categoria_subtotal",
            "missing",
            "blocking",
            "",
            "Subtotais de dashboard ainda nao sao emitidos pelo OV10.",
        ),
        "carteira_1": _gap(
            "carteira_1",
            "partial",
            "explained",
            "book_position_snapshots.book_id",
            (
                "Existe um book resolvido por posicao, "
                "mas nao as quatro colunas de carteiras do workbook."
            ),
        ),
        "carteira_2": _gap(
            "carteira_2",
            "missing",
            "blocking",
            "",
            "A matriz multi-carteira do workbook nao foi materializada como saida equivalente.",
        ),
        "carteira_3": _gap(
            "carteira_3",
            "missing",
            "blocking",
            "",
            "A matriz multi-carteira do workbook nao foi materializada como saida equivalente.",
        ),
        "carteira_4": _gap(
            "carteira_4",
            "missing",
            "blocking",
            "",
            "A matriz multi-carteira do workbook nao foi materializada como saida equivalente.",
        ),
    }
    gaps = tuple(support[item] for item in headers if item in support)
    summary = (
        "A aba `portfolio.` continua sendo o melhor análogo da visão canônica de posições, "
        "mas a paridade ainda é parcial: o núcleo patrimonial existe e faltam mercado, "
        "metadata de instrumento e a matriz multi-carteira do workbook."
    )
    metrics = {
        "reference_headers": headers,
        "open_position_snapshots": context.snapshot_count,
        "account_position_snapshots": context.account_snapshot_count,
        "resolved_book_ids": list(context.book_ids),
        "last_snapshot_date": context.last_snapshot_date,
        "positions_with_market_price": context.market_valuation.positions_with_market_price,
        "positions_with_display_name": context.market_valuation.positions_with_display_name,
        "positions_with_market_code": context.market_valuation.positions_with_market_code,
        "positions_with_avg_cost_in_price_currency": (
            context.market_valuation.positions_with_avg_cost_in_price_currency
        ),
    }
    return _build_area(
        area_id="portfolio_backend",
        sheet_name="portfolio.",
        reference_range="portfolio.!A3:Y3",
        summary=summary,
        gaps=gaps,
        metrics=metrics,
    )


def _build_cash_area(worksheet: _WorksheetLike, context: _RuntimeContext) -> ReferenceAreaReport:
    headers = [
        value
        for value in _clean_headers(
            worksheet=worksheet,
            row=4,
            start_column=12,
            end_column=23,
        )
        if value
    ]
    support = {
        "portfolio": _gap(
            "portfolio",
            "partial",
            "explained",
            "account_portfolio_assignment",
            "A atribuicao existe por conta, mas nao esta materializada em cada movimento de caixa.",
        ),
        "investidor": _gap(
            "investidor",
            "missing",
            "blocking",
            "",
            "Nao existe dimensao separada de investidor no ledger atual.",
        ),
        "corretora": _gap(
            "corretora",
            "matched",
            "expected",
            "cash_movement.account_id -> account_registry.account_name",
            "Cada movimento de caixa ja pertence a uma conta/corretora.",
        ),
        "saldo_inicial": _gap(
            "saldo_inicial",
            "missing",
            "blocking",
            "",
            "O OV10 ainda nao ingere saldo inicial nem caixa livre manual do workbook.",
        ),
        "moeda": _gap(
            "moeda",
            "matched",
            "expected",
            "cash_movement.currency",
            "Moeda ja existe em movimentos e saldos.",
        ),
        "rem/saq": _gap(
            "rem/saq",
            "missing",
            "blocking",
            "",
            "Remessas e saques manuais da aba `caixa` ainda nao entram no pipeline canonico.",
        ),
        "cre/deb": _gap(
            "cre/deb",
            "missing",
            "blocking",
            "",
            "Creditos e debitos manuais ainda nao entram no pipeline canonico.",
        ),
        "day tr": _gap(
            "day tr",
            "missing",
            "blocking",
            "",
            "O ledger atual nao separa bucket de day trade como o workbook.",
        ),
        "proventos": _gap(
            "proventos",
            "matched",
            "expected",
            "cash_movement[movement_type=dividend_receipt]",
            "Proventos ja entram como movimento de caixa derivado.",
        ),
        "fx caixa": _gap(
            "fx caixa",
            "missing",
            "blocking",
            "",
            "Nao existe camada de FX caixa e traducao monetaria equivalente ao workbook.",
        ),
        "saldo": _gap(
            "saldo",
            "matched",
            "expected",
            "cash_balance_snapshot.balance",
            "Saldo final por conta/moeda ja existe como snapshot persistivel.",
        ),
    }
    gaps = tuple(support[item] for item in headers if item in support)
    summary = (
        "A aba `caixa` deixa claro o maior gap atual do OV10: o ledger derivado cobre trades, "
        "proventos e eventos corporativos, mas ainda nao cobre saldo inicial, remessas/saques, "
        "creditos/debitos manuais e bucket dedicado de day trade."
    )
    metrics = {
        "reference_headers": headers,
        "cash_movements": context.cash_movement_count,
        "cash_balance_snapshots": context.cash_balance_snapshot_count,
    }
    return _build_area(
        area_id="cash_register",
        sheet_name="caixa",
        reference_range="caixa!L4:W4",
        summary=summary,
        gaps=gaps,
        metrics=metrics,
    )


def _build_allocation_area(
    worksheet: _WorksheetLike, context: _RuntimeContext
) -> ReferenceAreaReport:
    headers = [
        value
        for value in _clean_headers(
            worksheet=worksheet,
            row=4,
            start_column=2,
            end_column=16,
        )
        if value
    ]
    normalized_headers = [
        value
        for value in headers
        if value
        in {
            "BOOKS",
            "objetivo",
            "o que fazer",
            "ajuste (BRL)",
            "%",
            "VALOR DE MERCADO",
            "VALOR INVESTIDO",
            "FX. CAIXA",
            "LUCRO TOTAL",
        }
    ]
    support = {
        "BOOKS": _gap(
            "BOOKS",
            "matched",
            "expected",
            "book_registry + book_position_snapshots",
            "Os books já existem como entidade governada e como snapshots por posição.",
        ),
        "objetivo": _gap(
            "objetivo",
            "partial",
            "explained",
            "book_registry.target_weight",
            (
                "O peso alvo existe no contrato, "
                "mas ainda nao fecha a UX de balanceamento do workbook."
            ),
        ),
        "o que fazer": _gap(
            "o que fazer",
            "missing",
            "blocking",
            "",
            "Nao existe motor de recomendacao/rebalanceamento.",
        ),
        "ajuste (BRL)": _gap(
            "ajuste (BRL)",
            "missing",
            "blocking",
            "",
            "Nao existe calculo de ajuste monetario por book.",
        ),
        "%": _coverage_gap(
            "%",
            matched_count=context.market_valuation.books_with_current_weight,
            total_count=context.market_valuation.book_count,
            current_source="book_position_snapshots + latest market_snapshot/fx_snapshot",
            matched_note=(
                "A percentagem corrente ja pode ser derivada do valor de mercado "
                "persistido por book."
            ),
            partial_note="Existe cobertura parcial de percentagem corrente por valor de mercado.",
            missing_note=(
                "Nao existe cobertura persistida suficiente para percentagem corrente por book."
            ),
        ),
        "VALOR DE MERCADO": _coverage_gap(
            "VALOR DE MERCADO",
            matched_count=context.market_valuation.valued_book_count,
            total_count=context.market_valuation.book_count,
            current_source="book_position_snapshots + latest market_snapshot/fx_snapshot",
            matched_note="O valor de mercado por book ja pode ser agregado da camada persistida.",
            partial_note="Existe cobertura parcial de valor de mercado por book.",
            missing_note="Nao existe cobertura persistida de valor de mercado por book.",
        ),
        "VALOR INVESTIDO": _gap(
            "VALOR INVESTIDO",
            "matched",
            "expected",
            "book_position_snapshots.total_cost",
            "O valor investido por book pode ser agregado a partir do custo total.",
        ),
        "FX. CAIXA": _gap(
            "FX. CAIXA",
            "missing",
            "blocking",
            "",
            "Nao existe saida de caixa FX integrada ao book dashboard.",
        ),
        "LUCRO TOTAL": _coverage_gap(
            "LUCRO TOTAL",
            matched_count=context.market_valuation.books_with_profit_total,
            total_count=context.market_valuation.book_count,
            current_source="book_position_snapshots + latest market_snapshot/fx_snapshot",
            matched_note="O lucro total por book ja pode ser derivado da camada persistida.",
            partial_note="Existe cobertura parcial de lucro total por book.",
            missing_note="Nao existe cobertura persistida de lucro total por book.",
        ),
    }
    gaps = tuple(support[item] for item in normalized_headers if item in support)
    summary = (
        "A aba `alocação` ja tem fundacao contratual no OV10 via books e custo investido, "
        "mas ainda falta o bloco de mercado, ajuste e recomendacao que transforma isso em "
        "dashboard operacional de rebalanceamento."
    )
    metrics = {
        "reference_headers": normalized_headers,
        "book_snapshot_count": context.book_snapshot_count,
        "book_ids": list(context.book_ids),
        "valued_book_count": context.market_valuation.valued_book_count,
        "books_with_profit_total": context.market_valuation.books_with_profit_total,
        "books_with_current_weight": context.market_valuation.books_with_current_weight,
        "total_market_value_in_base_currency": (
            context.market_valuation.total_market_value_in_base_currency
        ),
    }
    return _build_area(
        area_id="allocation_dashboard",
        sheet_name="alocação",
        reference_range="alocação!B4:P4",
        summary=summary,
        gaps=gaps,
        metrics=metrics,
    )


def _build_income_area(
    *,
    worksheet_formulas: _WorksheetLike,
    worksheet_values: _WorksheetLike,
    context: _RuntimeContext,
) -> ReferenceAreaReport:
    month_end_dates = _extract_income_month_end_dates(worksheet_values)
    monthly_series = _compute_income_monthly_series(
        receipts=context.dividend_receipts,
        month_end_dates=month_end_dates,
    )
    formula_samples = [
        worksheet_formulas.cell(row=10, column=10).value,
        worksheet_formulas.cell(row=11, column=10).value,
        worksheet_formulas.cell(row=11, column=11).value,
    ]
    gaps = (
        _gap(
            "Ativo",
            "matched",
            "expected",
            "canonical_dividend_receipt.instrument_code",
            "Ativo do provento ja existe no fato canonico.",
        ),
        _gap(
            "Data",
            "matched",
            "expected",
            "canonical_dividend_receipt.received_date",
            "A data de recebimento ja existe no fato canonico.",
        ),
        _gap(
            "Evento",
            "matched",
            "expected",
            "canonical_dividend_receipt.dividend_type",
            "O tipo do evento de provento ja existe no fato canonico.",
        ),
        _gap(
            "Valor",
            "matched",
            "expected",
            "canonical_dividend_receipt.net_amount",
            "O valor liquido ja existe no fato canonico.",
        ),
        _gap(
            "Últimos 12 meses",
            "matched",
            "expected",
            "canonical_dividend_receipt grouped by workbook month windows",
            (
                "O OV10 ja consegue reproduzir os buckets mensais "
                "definidos pelas formulas do workbook."
            ),
        ),
    )
    summary = (
        "A aba `rendimentos` ja tem paridade conceitual forte: o OV10 possui os fatos canônicos "
        "de proventos e consegue reproduzir a janela mensal usada no painel de 12 meses."
    )
    metrics = {
        "reference_month_end_dates": month_end_dates[:12],
        "monthly_net_amounts": monthly_series,
        "formula_samples": formula_samples,
        "dividend_receipt_count": context.dividend_receipt_count,
    }
    return _build_area(
        area_id="income_dashboard",
        sheet_name="rendimentos",
        reference_range="rendimentos!G8:U10",
        summary=summary,
        gaps=gaps,
        metrics=metrics,
    )


def _extract_income_month_end_dates(worksheet: _WorksheetLike) -> list[date]:
    values: list[date] = []
    for column in range(10, 23):
        cell_value = worksheet.cell(row=8, column=column).value
        if isinstance(cell_value, datetime):
            values.append(cell_value.date())
        elif isinstance(cell_value, date):
            values.append(cell_value)
    if len(values) < 12:
        raise ValueError(
            "Reference workbook does not expose the expected month windows in `rendimentos`."
        )
    return values


def _compute_income_monthly_series(
    *,
    receipts: tuple[CanonicalDividendReceipt, ...],
    month_end_dates: list[date],
) -> list[dict[str, str]]:
    series: list[dict[str, str]] = []
    for index, month_end in enumerate(month_end_dates[:12]):
        lower_bound = (
            month_end_dates[index + 1]
            if index + 1 < len(month_end_dates)
            else _previous_month_end(month_end)
        )
        total = sum(
            (item.net_amount for item in receipts if lower_bound < item.received_date <= month_end),
            Decimal("0"),
        )
        series.append(
            {
                "month_end": month_end.isoformat(),
                "lower_bound_exclusive": lower_bound.isoformat(),
                "net_amount": str(total),
            }
        )
    return series


def _previous_month_end(value: date) -> date:
    if value.month == 1:
        return date(value.year - 1, 12, 31)
    return date(value.year, value.month, 1) - timedelta(days=1)


def _build_area(
    *,
    area_id: str,
    sheet_name: str,
    reference_range: str,
    summary: str,
    gaps: tuple[ReferenceGap, ...],
    metrics: dict[str, object],
) -> ReferenceAreaReport:
    overall_classification = "expected"
    if any(item.classification == "blocking" for item in gaps):
        overall_classification = "blocking"
    elif any(item.classification == "explained" for item in gaps):
        overall_classification = "explained"

    metrics_with_counts = {
        **metrics,
        "matched_fields": sum(1 for item in gaps if item.support_level == "matched"),
        "partial_fields": sum(1 for item in gaps if item.support_level == "partial"),
        "missing_fields": sum(1 for item in gaps if item.support_level == "missing"),
        "expected_fields": sum(1 for item in gaps if item.classification == "expected"),
        "explained_fields": sum(1 for item in gaps if item.classification == "explained"),
        "blocking_fields": sum(1 for item in gaps if item.classification == "blocking"),
    }
    return ReferenceAreaReport(
        area_id=area_id,
        sheet_name=sheet_name,
        reference_range=reference_range,
        overall_classification=overall_classification,
        summary=summary,
        gaps=gaps,
        metrics=metrics_with_counts,
    )


def _load_market_valuation(
    *,
    positions: list[PositionSnapshot],
    book_positions: list[BookPositionSnapshot],
    database_path: Path | None,
    snapshot_date: date | None,
) -> MarketValuationContext:
    if database_path is None or not database_path.exists() or not positions:
        return build_market_valuation_context(
            positions=positions,
            book_positions=book_positions,
            market_snapshots_by_code={},
            fx_snapshots_by_pair={},
        )

    with SQLiteOV10Store(database_path) as store:
        store.initialize()
        market_snapshots = store.fetch_latest_market_snapshots(
            canonical_codes=[item.instrument_code for item in positions],
            as_of_date=snapshot_date,
        )
        currencies = {item.currency for item in positions} | {
            str(snapshot.get("currency"))
            for snapshot in market_snapshots.values()
            if snapshot.get("currency")
        }
        fx_snapshots = store.fetch_latest_fx_snapshots(
            currencies=tuple(sorted(currencies)),
            as_of_date=snapshot_date,
        )

    return build_market_valuation_context(
        positions=positions,
        book_positions=book_positions,
        market_snapshots_by_code=market_snapshots,
        fx_snapshots_by_pair=fx_snapshots,
    )


def _coverage_gap(
    reference_field: str,
    *,
    matched_count: int,
    total_count: int,
    current_source: str,
    matched_note: str,
    partial_note: str,
    missing_note: str,
) -> ReferenceGap:
    if total_count <= 0:
        return _gap(reference_field, "missing", "blocking", "", missing_note)
    if matched_count >= total_count:
        return _gap(reference_field, "matched", "expected", current_source, matched_note)
    if matched_count > 0:
        return _gap(
            reference_field,
            "partial",
            "explained",
            current_source,
            f"{partial_note} Cobertura observada: {matched_count}/{total_count}.",
        )
    return _gap(reference_field, "missing", "blocking", "", missing_note)


def _gap(
    reference_field: str,
    support_level: str,
    classification: str,
    current_source: str,
    note: str,
) -> ReferenceGap:
    return ReferenceGap(
        reference_field=reference_field,
        support_level=support_level,
        classification=classification,
        current_source=current_source,
        note=note,
    )


def _clean_headers(
    *,
    worksheet: _WorksheetLike,
    row: int,
    start_column: int,
    end_column: int,
) -> list[str]:
    headers: list[str] = []
    for column in range(start_column, end_column + 1):
        value = worksheet.cell(row=row, column=column).value
        if value in (None, ""):
            continue
        headers.append(str(value).strip())
    return headers


def _serialize_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, date | datetime):
        return value.isoformat()
    if isinstance(value, tuple):
        return [_serialize_value(item) for item in value]
    if isinstance(value, list):
        return [_serialize_value(item) for item in value]
    if isinstance(value, dict):
        return {key: _serialize_value(item) for key, item in value.items()}
    return value
